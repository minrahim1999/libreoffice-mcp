"""Spreadsheet (xlsx) operations for LibreOffice MCP."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


def create_spreadsheet(output_path: str | Path) -> None:
    wb = Workbook()
    wb.save(str(output_path))


def add_sheet(file_path: str | Path, title: str) -> None:
    wb = load_workbook(str(file_path))
    if title in wb.sheetnames:
        raise ValueError(f"Sheet '{title}' already exists")
    wb.create_sheet(title=title)
    wb.save(str(file_path))


def set_cell_value(file_path: str | Path, sheet_title: str, cell: str, value: str) -> None:
    wb = load_workbook(str(file_path))
    if sheet_title not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_title}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet_title]
    ws[cell] = value
    wb.save(str(file_path))


def add_chart(
    file_path: str | Path,
    sheet_title: str,
    chart_type: str,
    data_range: str,
    title: str,
    position: str = "F2",
) -> None:
    wb = load_workbook(str(file_path))
    if sheet_title not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_title}' not found")
    ws = wb[sheet_title]

    # Parse data_range, e.g. "A1:D5"
    cells = ws[data_range]
    if not cells:
        raise ValueError(f"Invalid data_range: {data_range}")

    if chart_type == "bar":
        chart = BarChart()
        chart.type = "bar"
    elif chart_type == "column":
        chart = BarChart()
        chart.type = "col"
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    else:
        raise ValueError(f"Unsupported chart_type: {chart_type}")

    # cells is a tuple of tuples; determine top-left and bottom-right
    min_row = cells[0][0].row
    min_col = cells[0][0].column
    max_row = cells[-1][-1].row
    max_col = cells[-1][-1].column

    data_ref = Reference(ws, min_col=min_col, min_row=min_row + 1, max_col=max_col, max_row=max_row)
    cats_ref = Reference(ws, min_col=min_col, min_row=min_row + 1, max_col=max_col, max_row=max_row)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.title = title
    ws.add_chart(chart, position)
    wb.save(str(file_path))
